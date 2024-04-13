import openpyxl

wrkbk = openpyxl.Workbook()
sheet = wrkbk.active
title = sheet.title
print(f"Title of actibe sheet is: {title}")
sheet.title = "cui_sheet"
print(f"New title is {sheet.title}")

c1 = sheet.cell(row = 1, column = 1)
c1.value = "ali"
c2 = sheet.cell(row= 1 , column = 2)
c2.value = "alia"
c3 = sheet['A2']
c3.value = "sair"
c4 = sheet['B2']
c4.value = "saira"
c5 = sheet['D12']
c5.value = "bilal"
c6 = sheet['D11']
c6.value = "20"
c7 = sheet['D10']
c7.value = "2"
c8 = sheet['D13']
c8.value = c7.value+c6.value



wrkbk.save('cui_data_ds.xlsx')